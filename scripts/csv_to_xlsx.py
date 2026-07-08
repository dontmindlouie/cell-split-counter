"""Convert an events CSV to a formatted xlsx with sort/filter dropdowns.

Writes two sheets:
  - "events": raw 1-row-per-daughter dump, unchanged from earlier versions of this script.
  - "confirmed_splits": one row per unique split (deduplicated by parent_id + peak_frame,
    since every daughter row of the same split shares the same parent_id/peak_frame/centroid
    -- centroid is the PARENT's position at the split frame, see classify.py), filtered to
    claude_confidence > 0 (the documented "confirmed real" filter, see docs/output_schema.md),
    with an added is_fallback_review flag for the rare rows where review.py's error fallback
    fired (claude_confidence == tracker_persistence_score with no Claude notes/classification --
    these were never actually reviewed by Claude, just accepted at the tracker's own confidence).

Usage:
    python scripts/csv_to_xlsx.py data/output/events.csv
    python scripts/csv_to_xlsx.py data/output/events.csv --out data/output/events.xlsx
"""
import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _add_table_sheet(ws, table_name: str, rows: list[list[str]]) -> None:
    for row in rows:
        ws.append(row)

    if rows:
        n_rows, n_cols = len(rows), len(rows[0])
        last_col = get_column_letter(n_cols)
        table = Table(displayName=table_name, ref=f"A1:{last_col}{n_rows}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

        for col_idx in range(1, n_cols + 1):
            max_len = max((len(str(rows[r][col_idx - 1])) for r in range(n_rows)), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 60)
        ws.freeze_panes = "A2"


def _is_fallback_review(row: dict) -> bool:
    """True for review.py's silent error-fallback rows: accepted at the tracker's own
    confidence with no real Claude verdict behind it (see project memory 2026-07-06)."""
    try:
        conf = float(row["claude_confidence"])
        persist = float(row["tracker_persistence_score"])
    except (ValueError, KeyError):
        return False
    return (
        conf > 0
        and abs(conf - persist) < 1e-6
        and row.get("claude_notes", "") == ""
        and row.get("acd_division_type", "") == ""
    )


def _build_confirmed_splits(header: list[str], data_rows: list[list[str]]) -> list[list[str]]:
    fallback_col = "is_fallback_review"
    out_header = header + [fallback_col]
    seen: set[tuple[str, str]] = set()
    out_rows = [out_header]

    parent_idx = header.index("parent_id")
    frame_idx = header.index("peak_frame")
    conf_idx = header.index("claude_confidence")
    topology_idx = header.index("split_topology")

    for row in data_rows:
        # death rows are track ends, not splits -- this sheet is splits only.
        if row[topology_idx] not in ("normal_split", "multi_way_split"):
            continue
        try:
            confidence = float(row[conf_idx])
        except ValueError:
            continue
        if confidence <= 0:
            continue
        key = (row[parent_idx], row[frame_idx])
        if key in seen:
            continue
        seen.add(key)
        as_dict = dict(zip(header, row))
        out_rows.append(row + ["1" if _is_fallback_review(as_dict) else "0"])

    return out_rows


def _read_csv_rows(csv_path: Path) -> list[list[str]]:
    # CSVs written before src/output.py's explicit utf-8 fix (2026-07-06) were written in
    # whatever codepage Windows defaulted to (typically cp1252) -- fall back to that on
    # a decode error rather than failing outright on older run output still on disk.
    try:
        with open(csv_path, newline="", encoding="utf-8") as f:
            return list(csv.reader(f))
    except UnicodeDecodeError:
        with open(csv_path, newline="", encoding="cp1252") as f:
            return list(csv.reader(f))


def convert(csv_path: Path, xlsx_path: Path) -> None:
    rows = _read_csv_rows(csv_path)

    wb = Workbook()
    events_ws = wb.active
    events_ws.title = "events"
    _add_table_sheet(events_ws, "Events", rows)

    n_confirmed = 0
    if rows:
        confirmed_rows = _build_confirmed_splits(rows[0], rows[1:])
        confirmed_ws = wb.create_sheet("confirmed_splits")
        _add_table_sheet(confirmed_ws, "ConfirmedSplits", confirmed_rows)
        n_confirmed = len(confirmed_rows) - 1

    wb.active = 0  # "events" tab shown first on open

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"Wrote {xlsx_path} ({len(rows) - 1 if rows else 0} raw rows, {n_confirmed} unique confirmed splits)")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("csv_path", type=Path, help="Input CSV file")
    parser.add_argument(
        "--out", type=Path, default=None, help="Output xlsx path (default: same name, .xlsx)"
    )
    args = parser.parse_args()

    out_path = args.out or args.csv_path.with_name(args.csv_path.stem + "_formatted.xlsx")
    convert(args.csv_path, out_path)


if __name__ == "__main__":
    main()
